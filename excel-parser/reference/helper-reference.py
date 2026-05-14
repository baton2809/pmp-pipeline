"""
Sanity-test helper.py от SPEC excel-parser на реальном файле Натальи.
Цель: убедиться что код находит 28 задач, не 3.
"""
import openpyxl
import re
import json
import os
from datetime import datetime, timezone

CR_PATTERN = re.compile(r'(ASFC|CRSIGMA|OCRED|TIBDS|ASFS)-\d+')

def open_workbook(path):
    return openpyxl.load_workbook(path, data_only=True)

def get_sheet(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        return None
    return workbook[sheet_name]

def find_header_row(worksheet, max_search_rows=5):
    for row_idx in range(1, max_search_rows + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value and 'cr' in str(value).lower():
                return row_idx
    return None

def detect_columns(worksheet, header_row):
    name_patterns = {
        'cr': 'cr',
        'task': 'задача',
        'initiative': 'инициатива',
        'customer': 'заказчик',
        'analytics': 'аналитика',
        'development': 'разработка',
        'testing': 'тестирование',
    }
    columns = {key: [] for key in name_patterns}
    for col_idx in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(header_row, col_idx).value
        if cell_value is None:
            continue
        value_lower = str(cell_value).lower().strip()
        for key, pattern in name_patterns.items():
            if pattern in value_lower:
                columns[key].append(col_idx)
    return columns

def parse_cr_key(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    match = CR_PATTERN.search(text)
    return match.group(0) if match else None

def clean_text(value):
    if value is None:
        return ''
    text = str(value)
    text = text.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def extract_plan(row_idx, worksheet, columns):
    phases = {'analytics': None, 'development': None, 'testing': None}
    used_v2_for_any_phase = False
    
    for phase_key in phases:
        positions = columns.get(phase_key, [])
        for i, col_idx in enumerate(positions):
            cell_value = worksheet.cell(row_idx, col_idx).value
            if cell_value is not None and cell_value != '':
                try:
                    phases[phase_key] = float(cell_value)
                    if i > 0:
                        used_v2_for_any_phase = True
                except (ValueError, TypeError):
                    pass
    
    total = None
    any_filled = any(v is not None for v in phases.values())
    if any_filled:
        total = sum((v or 0) for v in phases.values())
    
    if not any_filled:
        source_version = 'none'
    elif used_v2_for_any_phase:
        source_version = 'v2'
    else:
        source_version = 'v1'
    
    return {
        'analytics': phases['analytics'],
        'development': phases['development'],
        'testing': phases['testing'],
        'total': total,
        'source_version': source_version,
    }

# === ПРОГОН ===
print("=== Тест helper.py на реальном файле Натальи ===\n")

wb = open_workbook("/mnt/user-data/uploads/Бэклог_и_цели__1_.xlsx")
ws = get_sheet(wb, "Q2_26_оценки_new_name")
print(f"Лист: {ws.title}, max_row={ws.max_row}, max_col={ws.max_column}\n")

header_row = find_header_row(ws)
print(f"Строка заголовков: {header_row}\n")

columns = detect_columns(ws, header_row)
print("Распознанные колонки:")
for key, positions in columns.items():
    if positions:
        col_letters = [openpyxl.utils.get_column_letter(p) for p in positions]
        print(f"  {key}: позиции {positions} (колонки {col_letters})")
print()

# Найти первую строку данных
first_data_row = None
cr_cols = columns.get('cr', [])
for row_idx in range(header_row + 1, ws.max_row + 1):
    for col_idx in cr_cols:
        value = ws.cell(row_idx, col_idx).value
        if value and CR_PATTERN.search(str(value)):
            first_data_row = row_idx
            break
    if first_data_row:
        break

print(f"Первая строка данных: {first_data_row}\n")

# Пройти и извлечь задачи
tasks = []
skipped_rows = []

for row_idx in range(first_data_row, ws.max_row + 1):
    cr_key = None
    for col_idx in cr_cols:
        value = ws.cell(row_idx, col_idx).value
        cr_key = parse_cr_key(value)
        if cr_key:
            break
    
    if not cr_key:
        # Проверим — может быть просто пустая строка или строка без CR (но с другими данными)
        # Чтобы не загромождать skipped_rows для пустых строк, проверим есть ли в строке хоть что-то
        has_any_value = any(
            ws.cell(row_idx, c).value not in (None, '') 
            for c in range(1, min(ws.max_column + 1, 21))
        )
        if has_any_value and cr_cols:
            # есть данные но нет CR-ключа в колонке cr
            cr_value = ws.cell(row_idx, cr_cols[0]).value if cr_cols else None
            if cr_value:  # что-то есть в колонке cr, но не CR-ключ
                skipped_rows.append({
                    'row': row_idx,
                    'reason': f'не удалось распознать CR в значении {str(cr_value)[:50]!r}'
                })
        continue
    
    # Извлекаем поля
    task_name = ''
    for c in columns.get('task', []):
        v = ws.cell(row_idx, c).value
        if v:
            task_name = clean_text(v)
            break
    
    initiative = ''
    for c in columns.get('initiative', []):
        v = ws.cell(row_idx, c).value
        if v:
            initiative = clean_text(v)
            break
    
    customer = None
    for c in columns.get('customer', []):
        v = ws.cell(row_idx, c).value
        if v:
            customer = clean_text(v)
            break
    
    plan = extract_plan(row_idx, ws, columns)
    
    tasks.append({
        'row': row_idx,
        'cr_key': cr_key,
        'task_name': task_name,
        'initiative': initiative,
        'customer': customer,
        'plan': plan,
    })

print(f"=== ИТОГ ===\n")
print(f"Извлечено задач: {len(tasks)}")
print(f"Пропущено строк (с мусором в cr): {len(skipped_rows)}")
print()

# Показать первые 5 задач
print("Первые 5 задач:")
for t in tasks[:5]:
    p = t['plan']
    plan_str = f"{p['analytics']}/{p['development']}/{p['testing']}={p['total']} ({p['source_version']})"
    print(f"  row {t['row']}: {t['cr_key']:18s} | план {plan_str:30s} | {t['task_name'][:40]!r}")

print(f"\nПоследние 3 задачи:")
for t in tasks[-3:]:
    p = t['plan']
    plan_str = f"{p['analytics']}/{p['development']}/{p['testing']}={p['total']} ({p['source_version']})"
    print(f"  row {t['row']}: {t['cr_key']:18s} | план {plan_str:30s} | {t['task_name'][:40]!r}")

# Покажу все уникальные CR
unique = list({t['cr_key'] for t in tasks})
print(f"\nУникальных CR-ключей: {len(unique)}")
print(f"Все ключи (отсортировано):")
for k in sorted(unique):
    print(f"  {k}")

# Распределение по проектам
from collections import Counter
projects = Counter(t['cr_key'].split('-')[0] for t in tasks)
print(f"\nПо проектам: {dict(projects)}")

# Скилл должен дать 28
print(f"\n{'='*50}")
expected = 28
got = len(tasks)
if got == expected:
    print(f"✅ УСПЕХ: найдено {got} задач (ожидалось {expected})")
else:
    print(f"❌ ПРОБЛЕМА: найдено {got} задач, ожидалось {expected}")
